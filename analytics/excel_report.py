import sys
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import sqlite3
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def get_traffic_status(speed):

    if speed < 15:
        return "Heavy Congestion"

    elif speed < 25:
        return "Moderate Traffic"

    else:
        return "Free Flow"


def format_sheet(worksheet):

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Style all cells
    for row in worksheet.iter_rows():

        for cell in row:

            cell.border = thin_border
            cell.alignment = center

    # Header row
    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font

    # Auto width
    for column_cells in worksheet.columns:

        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )

        worksheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 4

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = worksheet.dimensions


def main():

    project_root = Path(__file__).resolve().parent.parent

    db_path = project_root / "database" / "traffic.db"

    report_path = (
        project_root
        / "reports"
        / "Guntur_Traffic_Report.xlsx"
    )

    connection = sqlite3.connect(db_path)

    raw_df = pd.read_sql_query(
        "SELECT * FROM traffic_data",
        connection
    )

    if raw_df.empty:
        print("\nNo traffic data found in database. Run collector.py first.")
        connection.close()
        return

    summary_df = (
        raw_df.groupby("corridor_id")
        .agg(
            Average_Speed=("average_speed_kmph", "mean"),
            Average_Travel_Time=("duration_seconds", "mean"),
            Average_Distance=("distance_km", "mean"),
            Total_Observations=("corridor_id", "count")
        )
        .reset_index()
    )

    summary_df["Average_Speed"] = summary_df["Average_Speed"].round(2)
    summary_df["Average_Travel_Time"] = summary_df["Average_Travel_Time"].round(0)
    summary_df["Average_Distance"] = summary_df["Average_Distance"].round(2)

    summary_df["Traffic_Status"] = (
        summary_df["Average_Speed"]
        .apply(get_traffic_status)
    )

    from dashboard.congestion import CongestionAnalyzer
    rush_scale = CongestionAnalyzer.rush_hour_congestion_scale_0_10(raw_df)
    extra_info = CongestionAnalyzer.extra_time_per_50km(raw_df)
    net_info = CongestionAnalyzer.congested_road_network_pct(raw_df)
    rush_corr_df = CongestionAnalyzer.corridor_congestion_lengths(raw_df)

    dashboard_df = pd.DataFrame({
        "Metric": [
            "Total Records",
            "Average Speed (km/h)",
            "Average Travel Time (sec)",
            "Fastest Corridor",
            "Slowest Corridor",
            "Rush Hour Congestion Index (0-10)",
            "Extra Time per 50 km Travelled (min)",
            "Congested Road Network (%)",
            "Congested Network Length (km)"
        ],
        "Value": [
            len(raw_df),
            round(raw_df["average_speed_kmph"].mean(), 2),
            round(raw_df["duration_seconds"].mean(), 0),
            summary_df.loc[summary_df["Average_Speed"].idxmax(), "corridor_id"],
            summary_df.loc[summary_df["Average_Speed"].idxmin(), "corridor_id"],
            f"{rush_scale:.1f} / 10",
            f"+{extra_info['extra_time_min']:.1f} min",
            f"{net_info['congested_pct']:.1f}%",
            f"{net_info['congested_length_km']:.2f} km"
        ]
    })

    with pd.ExcelWriter(
        report_path,
        engine="openpyxl"
    ) as writer:

        dashboard_df.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

        summary_df.to_excel(
            writer,
            sheet_name="Corridor Summary",
            index=False
        )

        if not rush_corr_df.empty:
            rush_corr_df.to_excel(
                writer,
                sheet_name="Rush Hour Breakdown",
                index=False
            )

        raw_df.to_excel(
            writer,
            sheet_name="Raw Traffic Data",
            index=False
        )

        format_sheet(writer.sheets["Dashboard"])
        format_sheet(writer.sheets["Corridor Summary"])
        if "Rush Hour Breakdown" in writer.sheets:
            format_sheet(writer.sheets["Rush Hour Breakdown"])
        format_sheet(writer.sheets["Raw Traffic Data"])

    connection.close()

    print("\n" + "=" * 70)
    print("GUNTUR TRAFFIC REPORT GENERATED SUCCESSFULLY")
    print("=" * 70)


    print(f"\nSaved To:\n{report_path}")


if __name__ == "__main__":
    main()